from openpyxl import Workbook,load_workbook


"""
dic = {}
x = 1
while x < 6:
    size = ws[f"A{x}"].value
    price = ws[f"B{x}"].value
    dic.update({size :f"{price}"})
    x += 1
print(dic)
#Returns dictionary dic in form {'Tyres': 'Price', '135 40 14': '15', '155 55 16': '20', '195 55 16': '25', '185 35 18': '30'}
"""
def price_changer(file,qty,column):    #arg file and coloumn must be strings 
    """ Changes value of each cell in {coloumn} by {qty} in {file} (file AKA workbook).
        Parameters file and coloumn must be strings and qty must be an integer""" 
    "Can be made to work for non-price specific strings "                                    
    wb=load_workbook(file)
    ws = wb.active
    coloumn_ = ws[column]
    print(column)
    for cell in coloumn_:
        if type(cell.value) == str:
            cell.value = cell.value
        else:
            cell.value = cell.value + qty 
    wb.save(file)


def brand_changer(file,brand,column):  #args must be string
    """ Changes  the brand of tyres to {brand} in {coloumn} in {file}. arguments must be strings"""
    "Can be made to work for non-brand specific strings"                         
    wb=load_workbook(file)
    ws = wb.active
    column_ =ws[column]
    for cell in column_:
        if cell.value == "Brand":
            cell.value
        else:
            cell.value = f"{brand}"
    wb.save(file)
brand_changer("tyres2.xlsx","JAB","C")

def sheet_creator(workbook,name): 
    """ Creates an empty workshee, arguments must be strings"""
    wb = load_workbook(workbook)
    wb.create_sheet(name)
    wb.save(workbook)

def workbook_creator(name,title):
    """ Creates a workbook with  parameter {name} as its name,
    and a single sheet with the  parameter {title} as the name of the sheet,
    arguments must be strings"""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    wb.save(f"{name}.xlsx")
