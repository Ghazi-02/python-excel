from openpyxl import Workbook,load_workbook
import re

wb=load_workbook("tyres2.xlsx")

def minPriceSorter(file):
    """Takes in an excel file with worksheets(>=1) of tyres and returns a dictionary sorting
     through each worksheet to find the cheapest tyres.
     Can be made more generic
     re.sub("\s|\v|Z", "" ,size )"""
 
    dict = {}
    wb = load_workbook(file)
    for ws in wb:
        print(ws.title)
        column_=ws['A']
        x = 1
        while x <= len(column_):
            size = ws[f"A{x}"].value
            price = ws[f"B{x}"].value
            brand = ws[f"C{x}"].value
            if size in dict:
                if dict[size][0] <= price:
                    x += 1
                else:
                    dict[f"{size}"]=[price,f"{brand}"]
            elif size is None or price is None:
                x += 1
            else:   
                #dict.update({re.sub("\s|\v|Z", "" ,size ) :[price, f"{brand}" ]})  
                dict.update({size.replace(" ","").replace("Z",""):[price, f"{brand}" ]})
                x += 1
    print("DEBUG:",f"{dict}\n")
    print("DEBUG:",len(dict))
    return dict

def cheapestTyreTable(dictionary,file):
    """Turns a dictionary into an excel table. File should be the output destination of the table.
    Currenty used in conjunction with minPriceSorter().
    Can be made more generic"""
    y = 1
    while y < len(dictionary):
        wb = load_workbook(file)
        ws = wb.active
        for key in dictionary:
            ws[f"A{y}"].value = key
            ws[f"B{y}"].value = dictionary[key][0]
            ws[f"C{y}"].value = dictionary[key][1]
            wb.save(file)
            y+=1

cheapestTyreTable(minPriceSorter("tyres.xlsx"),"sorted.xlsx")